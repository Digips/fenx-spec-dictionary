"""Excel exporter.

Writes a :class:`~fenx_spec_dictionary.dictionary.Dictionary` to an
``.xlsx`` workbook using *openpyxl*.  The workbook contains the following
sheets:

* **All Elements** – every extracted element
* **Schemas** – schema / model properties only
* **Parameters** – path, query, header, and cookie parameters
* **Request Bodies** – request-body fields
* **Summary** – element counts per type
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fenx_spec_dictionary.dictionary import Dictionary

# Column definitions for the main element sheets
_ELEMENT_COLUMNS = [
    "Name",
    "Path",
    "Element Type",
    "Data Type",
    "Format",
    "Required",
    "Description",
    "Enum Values",
    "Example",
    "Parent",
    "Source",
    "Sources",
]

_HEADER_FILL = PatternFill(
    start_color="1F4E79", end_color="1F4E79", fill_type="solid"
)
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill(
    start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
)


def export(dictionary: Dictionary, output_path: str | Path) -> Path:
    """Export *dictionary* to an Excel workbook at *output_path*.

    Parameters
    ----------
    dictionary:
        The populated :class:`~fenx_spec_dictionary.dictionary.Dictionary`.
    output_path:
        Destination ``.xlsx`` file path.  The parent directory must exist.

    Returns
    -------
    Path
        The resolved path to the written workbook.
    """
    output_path = Path(output_path)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _write_elements_sheet(wb, "All Elements", dictionary.to_list())
    _write_elements_sheet(wb, "Schemas", dictionary.schemas())
    _write_elements_sheet(wb, "Parameters", dictionary.parameters())
    _write_elements_sheet(wb, "Request Bodies", dictionary.request_bodies())
    _write_summary_sheet(wb, dictionary.summary())

    wb.save(output_path)
    return output_path.resolve()


def _write_elements_sheet(
    wb: Workbook, title: str, rows: list[dict[str, Any]]
) -> None:
    ws = wb.create_sheet(title=title)

    # Header row
    for col_idx, col_name in enumerate(_ELEMENT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, entry in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(_ELEMENT_COLUMNS, start=1):
            value = entry.get(col_name, "")
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    # Auto-size columns (rough heuristic)
    for col_idx, col_name in enumerate(_ELEMENT_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[col_letter].width = max_len + 2

    # Enable auto-filter on the header row
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def _write_summary_sheet(wb: Workbook, summary: dict[str, int]) -> None:
    ws = wb.create_sheet(title="Summary")

    for col_idx, header in enumerate(["Element Type", "Count"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (element_type, count) in enumerate(
        sorted(summary.items()), start=2
    ):
        ws.cell(row=row_idx, column=1, value=element_type)
        ws.cell(row=row_idx, column=2, value=count)

    # Total row
    total_row = ws.max_row + 2
    total_cell = ws.cell(row=total_row, column=1, value="Total")
    total_cell.font = Font(bold=True)
    count_cell = ws.cell(row=total_row, column=2, value=sum(summary.values()))
    count_cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
