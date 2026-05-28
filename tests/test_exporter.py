"""Tests for the exporter module."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from fenx_spec_dictionary.dictionary import Dictionary
from fenx_spec_dictionary.exporter import export
from fenx_spec_dictionary.extractor import Element


def _make_dictionary() -> Dictionary:
    d = Dictionary()
    d.add_elements([
        Element(
            name="id",
            path="Entity.id",
            element_type="schema",
            data_type="string",
            format="uuid",
            required=True,
            description="Unique identifier",
            enum_values=[],
            example="abc-123",
            parent="Entity",
            source="spec.yaml",
        ),
        Element(
            name="status",
            path="Entity.status",
            element_type="schema",
            data_type="string",
            format="",
            required=False,
            description="Current status",
            enum_values=["Active", "Inactive"],
            example="Active",
            parent="Entity",
            source="spec.yaml",
        ),
        Element(
            name="limit",
            path="/entities#query.limit",
            element_type="parameter.query",
            data_type="integer",
            format="int32",
            required=False,
            description="Page size",
            enum_values=[],
            example="25",
            parent="listEntities",
            source="spec.yaml",
        ),
    ])
    return d


class TestExport:
    def test_creates_file(self, tmp_path):
        out = tmp_path / "output.xlsx"
        export(_make_dictionary(), out)
        assert out.exists()

    def test_expected_sheets(self, tmp_path):
        out = tmp_path / "output.xlsx"
        export(_make_dictionary(), out)
        wb = load_workbook(out)
        assert "All Elements" in wb.sheetnames
        assert "Schemas" in wb.sheetnames
        assert "Parameters" in wb.sheetnames
        assert "Request Bodies" in wb.sheetnames
        assert "Summary" in wb.sheetnames

    def test_all_elements_row_count(self, tmp_path):
        out = tmp_path / "output.xlsx"
        export(_make_dictionary(), out)
        wb = load_workbook(out)
        ws = wb["All Elements"]
        # header + 3 data rows
        assert ws.max_row == 4

    def test_schemas_sheet_filters_correctly(self, tmp_path):
        out = tmp_path / "output.xlsx"
        export(_make_dictionary(), out)
        wb = load_workbook(out)
        ws = wb["Schemas"]
        # header + 2 schema rows
        assert ws.max_row == 3

    def test_returns_resolved_path(self, tmp_path):
        out = tmp_path / "output.xlsx"
        result = export(_make_dictionary(), out)
        assert isinstance(result, Path)
        assert result.is_absolute()

    def test_empty_dictionary(self, tmp_path):
        out = tmp_path / "empty.xlsx"
        export(Dictionary(), out)
        assert out.exists()
