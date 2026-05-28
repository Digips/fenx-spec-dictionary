"""Integration tests for the end-to-end pipeline."""

from __future__ import annotations

from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from fenx_spec_dictionary.pipeline import run

_SAMPLE_SPEC = Path(__file__).parent.parent / "examples" / "sample_fenx_spec.yaml"


class TestPipelineIntegration:
    def test_run_with_sample_spec(self, tmp_path):
        result = run([str(_SAMPLE_SPEC)], output_path=tmp_path / "out.xlsx")
        assert result["element_count"] > 0
        assert Path(result["output"]).exists()

    def test_summary_contains_schemas(self, tmp_path):
        result = run([str(_SAMPLE_SPEC)], output_path=tmp_path / "out.xlsx")
        assert "schema" in result["summary"]

    def test_summary_contains_parameters(self, tmp_path):
        result = run([str(_SAMPLE_SPEC)], output_path=tmp_path / "out.xlsx")
        assert any(k.startswith("parameter") for k in result["summary"])

    def test_multiple_sources_deduplicated(self, tmp_path):
        # Running the same spec twice should not double the elements
        result_single = run(
            [str(_SAMPLE_SPEC)], output_path=tmp_path / "single.xlsx"
        )
        result_double = run(
            [str(_SAMPLE_SPEC), str(_SAMPLE_SPEC)],
            output_path=tmp_path / "double.xlsx",
        )
        assert result_single["element_count"] == result_double["element_count"]

    def test_excel_output_has_data(self, tmp_path):
        out = tmp_path / "out.xlsx"
        run([str(_SAMPLE_SPEC)], output_path=out)
        wb = load_workbook(out)
        ws = wb["All Elements"]
        # At least a header row + 1 data row
        assert ws.max_row >= 2
